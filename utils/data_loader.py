from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def load_json(path: str | Path) -> Any:
    p = Path(path)
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_excel_cases(path: str | Path, sheet_name: str = "Sheet1") -> list[dict[str, Any]]:
    """
    Excel 数据驱动约定：
    - 第一行是表头（key）
    - 从第二行开始是数据
    - 返回 list[dict]，用于 pytest 参数化
    """
    p = Path(path)
    wb = load_workbook(filename=str(p), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {p}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    cases: list[dict[str, Any]] = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        item = {headers[i]: r[i] for i in range(min(len(headers), len(r))) if headers[i]}
        cases.append(item)
    return cases

